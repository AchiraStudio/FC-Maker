import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading
import traceback
import pandas as pd
from openpyxl import load_workbook
import sv_ttk
from processor import TeamDataProcessor


class TeamProcessorApp:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("Team Data Processor")
        self.root.geometry("1200x800")
        self.root.minsize(900, 600)

        # ------------------------------------------------------------------
        # Initialise processor — show a friendly dialog on failure
        # ------------------------------------------------------------------
        try:
            self.processor = TeamDataProcessor()
            print("[app] Countries loaded:", self.processor.country_list[:5])
        except Exception as e:
            messagebox.showerror(
                "Initialisation Error",
                f"Failed to load data files:\n\n{e}\n\n"
                "Make sure the 'data/' folder is next to processor.py and "
                "contains all required JSON files."
            )
            self.root.destroy()
            return

        # ------------------------------------------------------------------
        # State variables
        # ------------------------------------------------------------------
        self.input_file    = tk.StringVar()
        self.output_file   = tk.StringVar()
        self.team_sheet    = tk.StringVar()
        self.player_sheet  = tk.StringVar()
        self.progress_value = tk.IntVar()
        self.progress_text  = tk.StringVar(value="Ready")
        self.teams_df      = None
        self.players_df    = None
        self.dark_mode     = False

        # Fonts
        self.default_font = ("Segoe UI", 10)
        self.title_font   = ("Segoe UI", 12, "bold")

        self._create_widgets()
        sv_ttk.set_theme("light")

    # ======================================================================
    # Widget construction
    # ======================================================================

    def _create_widgets(self):
        main = ttk.Frame(self.root)
        main.pack(fill="both", expand=True, padx=10, pady=10)

        # --- Header -------------------------------------------------------
        hdr = ttk.Frame(main)
        hdr.pack(fill="x", pady=(0, 8))
        ttk.Label(hdr, text="Team Data Processor", font=self.title_font).pack(side="left")
        self.theme_btn = ttk.Button(hdr, text="☀️", command=self._toggle_theme, width=3)
        self.theme_btn.pack(side="right", padx=5)

        # --- Input file ---------------------------------------------------
        in_frame = ttk.LabelFrame(main, text="Input File", padding=8)
        in_frame.pack(fill="x", pady=4)
        ttk.Label(in_frame, text="Excel File:").grid(row=0, column=0, sticky="e", padx=(0,4))
        ttk.Entry(in_frame, textvariable=self.input_file, width=70).grid(row=0, column=1, padx=4)
        ttk.Button(in_frame, text="Browse", command=self._select_input_file).grid(row=0, column=2)

        # --- Sheet selection ----------------------------------------------
        sh_frame = ttk.LabelFrame(main, text="Sheet Selection", padding=8)
        sh_frame.pack(fill="x", pady=4)
        ttk.Label(sh_frame, text="Team Sheet:").grid(row=0, column=0, sticky="e", padx=(0,4))
        self.team_sheet_cb = ttk.Combobox(sh_frame, textvariable=self.team_sheet,
                                          state="readonly", width=25)
        self.team_sheet_cb.grid(row=0, column=1, padx=4, sticky="w")
        self.team_sheet_cb.bind("<<ComboboxSelected>>", lambda _: self._reload_data())

        ttk.Label(sh_frame, text="Player Sheet:").grid(row=0, column=2, sticky="e", padx=(12,4))
        self.player_sheet_cb = ttk.Combobox(sh_frame, textvariable=self.player_sheet,
                                            state="readonly", width=25)
        self.player_sheet_cb.grid(row=0, column=3, padx=4, sticky="w")
        self.player_sheet_cb.bind("<<ComboboxSelected>>", lambda _: self._reload_data())

        # --- Output file --------------------------------------------------
        out_frame = ttk.LabelFrame(main, text="Output File", padding=8)
        out_frame.pack(fill="x", pady=4)
        ttk.Label(out_frame, text="Output File:").grid(row=0, column=0, sticky="e", padx=(0,4))
        ttk.Entry(out_frame, textvariable=self.output_file, width=70).grid(row=0, column=1, padx=4)
        ttk.Button(out_frame, text="Browse", command=self._select_output_file).grid(row=0, column=2)

        # --- Progress bar -------------------------------------------------
        prog_frame = ttk.LabelFrame(main, text="Progress", padding=8)
        prog_frame.pack(fill="x", pady=4)
        ttk.Label(prog_frame, textvariable=self.progress_text).pack()
        self.progress_bar = ttk.Progressbar(prog_frame, variable=self.progress_value,
                                            maximum=100, length=400)
        self.progress_bar.pack(fill="x")

        # --- Notebook -----------------------------------------------------
        nb_frame = ttk.Frame(main)
        nb_frame.pack(fill="both", expand=True, pady=4)
        self.notebook = ttk.Notebook(nb_frame)
        self.notebook.pack(fill="both", expand=True)

        self._build_teams_tab()
        self._build_players_tab()
        self._build_managers_tab()
        self._build_log_tab()

        # --- Buttons ------------------------------------------------------
        btn_frame = ttk.Frame(main)
        btn_frame.pack(pady=8)
        self.process_btn = ttk.Button(btn_frame, text="START PROCESS",
                                      command=self._start_processing,
                                      style='Accent.TButton')
        self.process_btn.pack(side="left", padx=5, ipadx=20)
        ttk.Button(btn_frame, text="Exit", command=self.root.quit).pack(side="left", padx=5)

        self.style = ttk.Style()
        self.style.configure('Accent.TButton', font=('Segoe UI', 10, 'bold'))

    def _build_tree(self, parent, columns: list) -> ttk.Treeview:
        """Helper — builds a Treeview with a vertical scrollbar."""
        frame = ttk.Frame(parent)
        frame.pack(fill="both", expand=True)
        tree = ttk.Treeview(frame, columns=columns, show="headings")
        for col in columns:
            tree.heading(col, text=col.capitalize())
            tree.column(col, width=120, minwidth=60)
        sb = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=sb.set)
        tree.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        return tree

    def _build_teams_tab(self):
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="Teams")
        self.teams_tree = self._build_tree(tab, ["teamid", "teamname"])
        self.teams_count = ttk.Label(tab, text="Total Teams: 0")
        self.teams_count.pack(side="bottom", anchor="w", padx=4)

    def _build_players_tab(self):
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="Players")
        self.players_tree = self._build_tree(
            tab, ["playerid", "playername", "team", "ovr", "position"])
        self.players_count = ttk.Label(tab, text="Total Players: 0")
        self.players_count.pack(side="bottom", anchor="w", padx=4)

    def _build_managers_tab(self):
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="Managers")
        self.managers_tree = self._build_tree(
            tab, ["managerid", "firstname", "lastname", "teamid"])
        self.managers_count = ttk.Label(tab, text="Total Managers: 0")
        self.managers_count.pack(side="bottom", anchor="w", padx=4)

    def _build_log_tab(self):
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="Log")
        self.log_area = scrolledtext.ScrolledText(
            tab, wrap=tk.WORD, font=("Consolas", 9))
        self.log_area.pack(fill="both", expand=True)

    # ======================================================================
    # Theme toggle
    # ======================================================================

    def _toggle_theme(self):
        self.dark_mode = not self.dark_mode
        sv_ttk.set_theme("dark" if self.dark_mode else "light")
        self.theme_btn.config(text="🌙" if self.dark_mode else "☀️")
        bg = '#2d2d2d' if self.dark_mode else 'white'
        fg = 'white'  if self.dark_mode else 'black'
        self.log_area.config(bg=bg, fg=fg, insertbackground=fg)

    # ======================================================================
    # File dialogs
    # ======================================================================

    def _select_input_file(self):
        path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx *.xls")])
        if path:
            self.input_file.set(path)
            self._load_sheet_names(path)

    def _select_output_file(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")])
        if path:
            self.output_file.set(path)

    # ======================================================================
    # Data loading
    # ======================================================================

    def _load_sheet_names(self, path: str):
        try:
            wb = load_workbook(filename=path, read_only=True)
            sheets = wb.sheetnames
            wb.close()
            self.team_sheet_cb['values']   = sheets
            self.player_sheet_cb['values'] = sheets
            self.team_sheet.set(sheets[0])
            self.player_sheet.set(sheets[1] if len(sheets) > 1 else sheets[0])
            self._log(f"Opened: {path}\nSheets: {', '.join(sheets)}")
            self._reload_data()
        except Exception as e:
            messagebox.showerror("File Error", f"Could not open file:\n{e}")

    def _reload_data(self):
        """Re-read both sheets and refresh the treeviews."""
        if not self.input_file.get():
            return
        try:
            # Teams
            self.teams_df = pd.read_excel(
                self.input_file.get(), sheet_name=self.team_sheet.get())
            self.teams_df.columns = [c.strip() for c in self.teams_df.columns]

            self.teams_tree.delete(*self.teams_tree.get_children())
            for _, row in self.teams_df.iterrows():
                self.teams_tree.insert("", "end",
                    values=(row.get('teamid', ''), row.get('teamname', '')))
            self.teams_count.config(text=f"Total Teams: {len(self.teams_df)}")

            # Players
            self.players_df = pd.read_excel(
                self.input_file.get(), sheet_name=self.player_sheet.get())
            self.players_df.columns = [c.strip() for c in self.players_df.columns]

            # Resolve name column flexibly
            name_col = next((c for c in ['Name', 'playername', 'Firstname']
                             if c in self.players_df.columns), None)
            pos_col  = next((c for c in ['Position', 'Position1']
                             if c in self.players_df.columns), None)

            self.players_tree.delete(*self.players_tree.get_children())
            for _, row in self.players_df.iterrows():
                self.players_tree.insert("", "end", values=(
                    row.get('playerid', ''),
                    row[name_col] if name_col else '',
                    row.get('Team', ''),
                    row.get('OVR', ''),
                    row[pos_col] if pos_col else '',
                ))
            self.players_count.config(text=f"Total Players: {len(self.players_df)}")

            # Clear stale managers
            self.managers_tree.delete(*self.managers_tree.get_children())
            self.managers_count.config(text="Total Managers: 0")

            self._log(f"Loaded {len(self.teams_df)} teams and "
                      f"{len(self.players_df)} players.")
        except Exception as e:
            self._log(f"[ERROR] loading data: {e}\n{traceback.format_exc()}")
            messagebox.showerror("Load Error", f"Failed to load sheets:\n{e}")

    # ======================================================================
    # Processing
    # ======================================================================

    def _start_processing(self):
        if self.teams_df is None or self.players_df is None:
            messagebox.showerror("Error", "Please load an input file first.")
            return
        if not self.output_file.get():
            messagebox.showerror("Error", "Please select an output file.")
            return

        self.process_btn.config(state='disabled', text="Processing…")
        self.log_area.delete("1.0", tk.END)
        threading.Thread(target=self._process_worker, daemon=True).start()

    def _process_worker(self):
        """Runs in a background thread — never touches widgets directly."""
        try:
            self._safe_update_progress(10, "Validating data…")
            self._safe_log(f"Teams columns  : {self.teams_df.columns.tolist()}")
            self._safe_log(f"Players columns: {self.players_df.columns.tolist()}")

            self._safe_update_progress(30, "Processing teams…")
            processed = self.processor.process_data(self.teams_df, self.players_df)

            self._safe_update_progress(70, "Updating display…")
            # Schedule GUI updates on the main thread
            self.root.after(0, lambda: self._populate_managers(processed))

            self._safe_update_progress(80, "Saving to Excel…")
            self._save_to_excel(processed, self.output_file.get())

            self._safe_update_progress(100, "Done!")
            self._safe_log("Processing completed successfully!")
            self.root.after(0, lambda: messagebox.showinfo(
                "Success", f"Saved to:\n{self.output_file.get()}"))

        except Exception as e:
            err = f"[ERROR] {e}\n{traceback.format_exc()}"
            self._safe_log(err)
            self.root.after(0, lambda: messagebox.showerror(
                "Processing Error", f"Processing failed:\n{e}"))
        finally:
            self.root.after(0, self._reset_process_btn)
            self.root.after(0, lambda: self._safe_update_progress(0, "Ready"))

    def _populate_managers(self, processed: dict):
        """Called on the main thread — safe to touch widgets."""
        if 'managers' not in processed:
            return
        df = processed['managers']
        self.managers_tree.delete(*self.managers_tree.get_children())
        for _, row in df.iterrows():
            self.managers_tree.insert("", "end", values=(
                row.get('managerid', ''),
                row.get('firstname', ''),
                row.get('lastname', ''),
                row.get('teamid', ''),
            ))
        self.managers_count.config(text=f"Total Managers: {len(df)}")

    def _reset_process_btn(self):
        self.process_btn.config(state='normal', text="START PROCESS")

    # ======================================================================
    # Excel export
    # ======================================================================

    def _save_to_excel(self, data: dict, path: str):
        with pd.ExcelWriter(path, engine='openpyxl') as writer:
            for sheet_name, df in data.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        self._safe_log(f"Saved → {path}")

    # ======================================================================
    # Thread-safe helpers
    # ======================================================================

    def _safe_log(self, message: str):
        """Log from any thread."""
        self.root.after(0, lambda m=message: self._log(m))

    def _safe_update_progress(self, value: int, text: str):
        self.root.after(0, lambda v=value, t=text: self._update_progress(v, t))

    def _log(self, message: str):
        self.log_area.insert(tk.END, message + "\n")
        self.log_area.see(tk.END)

    def _update_progress(self, value: int, text: str):
        self.progress_value.set(value)
        self.progress_text.set(text)


# ==========================================================================
# Entry point
# ==========================================================================

if __name__ == "__main__":
    root = tk.Tk()
    app = TeamProcessorApp(root)
    root.mainloop()
